#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
重复文件检测脚本 - 带图片预览功能
依赖库：pip install openpyxl pillow
"""

import os
import hashlib
import sys
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage


def calculate_md5(file_path):
    """计算文件的MD5值"""
    md5_hash = hashlib.md5()
    try:
        with open(file_path, 'rb') as f:
            # 分块读取文件，避免大文件占用过多内存
            for chunk in iter(lambda: f.read(4096), b''):
                md5_hash.update(chunk)
        return md5_hash.hexdigest()
    except Exception as e:
        print(f"无法读取文件 {file_path}: {e}")
        return None


def is_image_file(file_path):
    """检查文件是否为图片格式"""
    image_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp', '.ico'}
    return Path(file_path).suffix.lower() in image_extensions


def create_thumbnail(image_path, max_size=(150, 150)):
    """创建缩略图"""
    try:
        with PILImage.open(image_path) as img:
            # 转换为RGB模式（处理RGBA等格式）
            if img.mode in ('RGBA', 'LA', 'P'):
                background = PILImage.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                img = background
            
            # 创建缩略图
            img.thumbnail(max_size, PILImage.Resampling.LANCZOS)
            
            # 保存临时文件
            temp_path = f"temp_thumb_{os.path.basename(image_path)}.png"
            img.save(temp_path, 'PNG')
            return temp_path
    except Exception as e:
        print(f"无法创建缩略图 {image_path}: {e}")
        return None


def find_duplicate_files(directory):
    """查找目录下所有重复的文件"""
    if not os.path.exists(directory):
        print(f"错误：目录 '{directory}' 不存在")
        sys.exit(1)
    
    if not os.path.isdir(directory):
        print(f"错误：'{directory}' 不是一个有效的目录")
        sys.exit(1)
    
    # 使用字典存储MD5值和对应的文件路径列表
    md5_dict = defaultdict(list)
    
    print(f"正在扫描目录: {directory}")
    file_count = 0
    
    # 遍历目录下的所有文件
    for root, dirs, files in os.walk(directory):
        for filename in files:
            file_path = os.path.join(root, filename)
            file_count += 1
            
            if file_count % 100 == 0:
                print(f"已扫描 {file_count} 个文件...")
            
            md5_value = calculate_md5(file_path)
            if md5_value:
                md5_dict[md5_value].append(file_path)
    
    print(f"扫描完成，共扫描 {file_count} 个文件")
    
    # 筛选出重复的文件（MD5值相同的文件数量大于1）
    duplicate_files = {md5: paths for md5, paths in md5_dict.items() if len(paths) > 1}
    
    return duplicate_files


def export_to_excel(duplicate_files, output_file):
    """将重复文件信息导出到Excel"""
    if not duplicate_files:
        print("没有发现重复的文件")
        return
    
    # 确保输出目录存在
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"创建输出目录: {output_dir}")
    
    # 创建工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "重复文件"
    
    # 设置表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # 写入表头
    headers = ["组号", "预览图", "MD5值", "文件路径", "文件大小(字节)", "重复文件数量"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # 写入数据
    row = 2
    group_num = 1
    temp_files = []  # 存储临时文件路径，稍后删除
    
    for md5, paths in duplicate_files.items():
        duplicate_count = len(paths)  # 计算该组重复文件的数量
        start_row = row  # 记录该组的起始行
        
        # 检查第一个文件是否为图片，用于预览图
        first_image_path = None
        for path in paths:
            if is_image_file(path):
                first_image_path = path
                break
        
        for path in paths:
            try:
                file_size = os.path.getsize(path)
            except:
                file_size = "无法获取"
            
            # 只写入MD5值、文件路径、文件大小（组号和预览图稍后合并处理）
            ws.cell(row=row, column=3, value=md5)
            ws.cell(row=row, column=4, value=path)
            ws.cell(row=row, column=5, value=file_size)
            
            row += 1
        
        # 设置该组的行高（如果有图片的话）
        if first_image_path:
            # 计算合适的行高：图片高度除以重复文件数量，但不低于最小行高
            image_height = 120  # 图片高度（磅）
            min_row_height = 20  # 最小行高（磅），确保文字可读
            group_row_count = len(paths)
            
            # 计算每行应该的高度
            calculated_height = image_height / group_row_count
            final_height = max(calculated_height, min_row_height)
            
            # 为该组的所有行设置统一的行高
            for group_row in range(start_row, row):
                ws.row_dimensions[group_row].height = final_height
        
        end_row = row - 1
        
        # 合并该组的"组号"单元格并设置值
        if start_row == end_row:
            # 只有一行，不需要合并
            cell = ws.cell(row=start_row, column=1, value=group_num)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            # 多行，合并单元格
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            cell = ws.cell(row=start_row, column=1, value=group_num)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 合并该组的"预览图"单元格并插入图片（如果有图片文件）
        if first_image_path:
            try:
                thumb_path = create_thumbnail(first_image_path)
                if thumb_path:
                    if start_row == end_row:
                        # 只有一行，不需要合并
                        pass  # 单元格已存在，直接插入图片即可
                    else:
                        # 多行，合并单元格
                        ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                    
                    img = ExcelImage(thumb_path)
                    # 调整图片大小以适应单元格
                    img.width = 150
                    img.height = 150
                    
                    # 插入图片到预览图列（第2列）的起始行
                    ws.add_image(img, f'B{start_row}')
                    
                    # 将临时文件路径添加到列表，稍后删除
                    temp_files.append(thumb_path)
            except Exception as e:
                print(f"插入预览图失败 {first_image_path}: {e}")
        else:
            # 没有图片文件，只合并预览图单元格（保持空白）
            if start_row != end_row:
                ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
        
        # 为预览图列设置居中对齐（无论是否有图片）
        if start_row == end_row:
            # 只有一行
            preview_cell = ws.cell(row=start_row, column=2)
            preview_cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            # 多行，已合并的单元格
            preview_cell = ws.cell(row=start_row, column=2)
            preview_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 合并该组的"重复文件数量"单元格并设置值
        if start_row == end_row:
            # 只有一行，不需要合并
            cell = ws.cell(row=start_row, column=6, value=duplicate_count)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            # 多行，合并单元格
            ws.merge_cells(start_row=start_row, start_column=6, end_row=end_row, end_column=6)
            cell = ws.cell(row=start_row, column=6, value=duplicate_count)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        group_num += 1
    
    # 调整列宽（增大以便完整显示内容）
    ws.column_dimensions['A'].width = 12   # 组号
    ws.column_dimensions['B'].width = 22   # 预览图列
    ws.column_dimensions['C'].width = 40   # MD5值需要32个字符
    ws.column_dimensions['D'].width = 120  # 文件路径需要更宽
    ws.column_dimensions['E'].width = 20   # 文件大小
    ws.column_dimensions['F'].width = 18   # 重复文件数量
    
    # 保存文件
    wb.save(output_file)
    
    # 清理临时缩略图文件
    for temp_file in temp_files:
        try:
            os.remove(temp_file)
        except:
            pass
    
    print(f"\n结果已保存到: {output_file}")
    print(f"共发现 {len(duplicate_files)} 组重复文件")
    
    # 输出统计信息
    total_duplicates = sum(len(paths) for paths in duplicate_files.values())
    print(f"重复文件总数: {total_duplicates} 个")
    
    # 打印详细信息
    print("\n重复文件详情：")
    group_num = 1
    for md5, paths in duplicate_files.items():
        print(f"\n第 {group_num} 组 (MD5: {md5}):")
        for path in paths:
            try:
                size = os.path.getsize(path)
                print(f"  - {path} ({size} 字节)")
            except:
                print(f"  - {path}")
        group_num += 1


def main():
    """主函数"""
    if len(sys.argv) < 2:
        print("使用方法: python find_duplicate_files.py <目录路径> [输出路径]")
        print("示例1: python find_duplicate_files.py C:\\MyFolder")
        print("示例2: python find_duplicate_files.py C:\\MyFolder D:\\output")
        sys.exit(1)
    
    directory = sys.argv[1]
    
    # 获取目录名称（用于生成文件名）
    dir_name = os.path.basename(os.path.abspath(directory))
    excel_filename = f"same_file_in_{dir_name}.xlsx"
    
    # 确定输出路径
    if len(sys.argv) >= 3:
        # 如果提供了输出路径参数，在该路径下创建文件
        output_path = sys.argv[2]
        output_file = os.path.join(output_path, excel_filename)
        print(f"输出文件将保存到指定路径: {output_file}")
    else:
        # 如果没有提供输出路径，在指定目录的同级目录下创建文件
        parent_dir = os.path.dirname(os.path.abspath(directory))
        output_file = os.path.join(parent_dir, excel_filename)
        print(f"输出文件将保存到同级目录: {output_file}")
    
    # 查找重复文件
    duplicate_files = find_duplicate_files(directory)
    
    # 导出到Excel
    export_to_excel(duplicate_files, output_file)


if __name__ == "__main__":
    main()