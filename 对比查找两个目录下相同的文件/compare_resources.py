#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
游戏资源文件对比工具
通过MD5值匹配两个文件夹中内容相同但名称不同的文件
"""

import os
import hashlib
import argparse
from pathlib import Path
from typing import Dict, List, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def calculate_md5(file_path: str) -> str:
    """
    计算文件的MD5值
    
    Args:
        file_path: 文件路径
        
    Returns:
        文件的MD5值（十六进制字符串）
    """
    md5_hash = hashlib.md5()
    try:
        with open(file_path, 'rb') as f:
            # 分块读取文件，避免大文件占用过多内存
            for chunk in iter(lambda: f.read(4096), b''):
                md5_hash.update(chunk)
        return md5_hash.hexdigest()
    except Exception as e:
        print(f"计算MD5失败: {file_path}, 错误: {e}")
        return ""


def scan_directory(directory: str) -> Dict[str, List[str]]:
    """
    扫描目录，计算所有文件的MD5值
    
    Args:
        directory: 要扫描的目录路径
        
    Returns:
        字典，key为MD5值，value为文件路径列表（可能多个文件有相同MD5）
    """
    md5_dict = {}
    directory = os.path.abspath(directory)
    
    print(f"\n正在扫描目录: {directory}")
    file_count = 0
    
    for root, dirs, files in os.walk(directory):
        for filename in files:
            file_path = os.path.join(root, filename)
            md5_value = calculate_md5(file_path)
            
            if md5_value:
                if md5_value not in md5_dict:
                    md5_dict[md5_value] = []
                md5_dict[md5_value].append(file_path)
                file_count += 1
                
                if file_count % 100 == 0:
                    print(f"  已处理 {file_count} 个文件...")
    
    print(f"  完成！共处理 {file_count} 个文件，{len(md5_dict)} 个唯一MD5值")
    return md5_dict


def compare_directories(dir1: str, dir2: str) -> List[Tuple[str, List[str], List[str]]]:
    """
    对比两个目录中MD5相同的文件
    
    Args:
        dir1: 第一个目录路径（国内版本）
        dir2: 第二个目录路径（国外版本）
        
    Returns:
        匹配结果列表，每个元素为 (md5值, dir1中的文件列表, dir2中的文件列表)
    """
    print("\n" + "="*60)
    print("开始扫描文件...")
    print("="*60)
    
    md5_dict1 = scan_directory(dir1)
    md5_dict2 = scan_directory(dir2)
    
    # 找到共同的MD5值
    common_md5 = set(md5_dict1.keys()) & set(md5_dict2.keys())
    
    print(f"\n找到 {len(common_md5)} 个MD5值匹配的文件组")
    
    # 构建结果列表
    results = []
    for md5_value in sorted(common_md5):
        results.append((md5_value, md5_dict1[md5_value], md5_dict2[md5_value]))
    
    return results


def get_relative_path(file_path: str, base_dir: str) -> str:
    """
    获取相对路径
    
    Args:
        file_path: 文件绝对路径
        base_dir: 基础目录
        
    Returns:
        相对路径
    """
    try:
        return os.path.relpath(file_path, base_dir)
    except:
        return file_path


def export_to_excel(results: List[Tuple[str, List[str], List[str]]], 
                   output_path: str, 
                   dir1: str, 
                   dir2: str):
    """
    将对比结果导出到Excel文件
    
    Args:
        results: 对比结果列表
        output_path: 输出Excel文件路径
        dir1: 第一个目录路径
        dir2: 第二个目录路径
    """
    print(f"\n正在生成Excel文件: {output_path}")
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资源对比结果"
    
    # 设置表头样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入表头
    headers = ["序号", "MD5值", f"文件路径1 ({os.path.basename(dir1)})", 
               f"文件路径2 ({os.path.basename(dir2)})", "文件大小(字节)"]
    ws.append(headers)
    
    # 设置表头样式
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 写入数据
    row_num = 2
    for idx, (md5_value, files1, files2) in enumerate(results, 1):
        # 获取文件大小（使用第一个文件）
        file_size = os.path.getsize(files1[0]) if files1 else 0
        
        # 如果有多个文件对应，每个组合占一行
        max_count = max(len(files1), len(files2))
        
        for i in range(max_count):
            file1 = get_relative_path(files1[i], dir1) if i < len(files1) else ""
            file2 = get_relative_path(files2[i], dir2) if i < len(files2) else ""
            
            row_data = [
                idx if i == 0 else "",  # 序号只在第一行显示
                md5_value if i == 0 else "",  # MD5只在第一行显示
                file1,
                file2,
                file_size if i == 0 else ""  # 文件大小只在第一行显示
            ]
            ws.append(row_data)
            
            # 设置对齐方式
            for col_num in range(1, 6):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            row_num += 1
    
    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 15
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存文件
    wb.save(output_path)
    print(f"Excel文件生成成功！共 {len(results)} 组匹配文件")


def main():
    """
    主函数
    """
    parser = argparse.ArgumentParser(
        description='游戏资源文件对比工具 - 通过MD5值匹配两个文件夹中的对应文件',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  python compare_resources.py D:/assets/cn D:/assets/en D:/output/contrast.xlsx
  python compare_resources.py "C:/Game/Resources/CN" "C:/Game/Resources/EN" "C:/output/result.xlsx"
        """
    )
    
    parser.add_argument('dir1', type=str, help='第一个文件夹路径（国内版本）')
    parser.add_argument('dir2', type=str, help='第二个文件夹路径（国外版本）')
    parser.add_argument('output', type=str, help='输出Excel文件路径（例如: contrast.xlsx）')
    
    args = parser.parse_args()
    
    # 验证输入目录
    if not os.path.isdir(args.dir1):
        print(f"错误: 目录不存在: {args.dir1}")
        return
    
    if not os.path.isdir(args.dir2):
        print(f"错误: 目录不存在: {args.dir2}")
        return
    
    # 确保输出目录存在
    output_dir = os.path.dirname(args.output)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"创建输出目录: {output_dir}")
    
    # 确保输出文件有.xlsx扩展名
    if not args.output.lower().endswith('.xlsx'):
        args.output += '.xlsx'
    
    print("="*60)
    print("游戏资源文件对比工具")
    print("="*60)
    print(f"目录1: {args.dir1}")
    print(f"目录2: {args.dir2}")
    print(f"输出文件: {args.output}")
    
    # 执行对比
    results = compare_directories(args.dir1, args.dir2)
    
    # 导出到Excel
    if results:
        export_to_excel(results, args.output, args.dir1, args.dir2)
        print("\n处理完成！")
    else:
        print("\n警告: 没有找到任何匹配的文件")


if __name__ == '__main__':
    main()

