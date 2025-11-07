#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
游戏资源文件对比工具（带图片预览）
通过MD5值匹配两个文件夹中内容相同但名称不同的文件，并在Excel中显示图片预览
"""

import os
import hashlib
import argparse
from pathlib import Path
from typing import Dict, List, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from PIL import Image
import io
import tempfile


# 支持的图片格式
IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff', '.tif', '.webp'}


def calculate_md5(file_path: str) -> str:
    """
    计算文件的MD5值
    
    Args:
        file_path: 文件路径
        
    Returns:
        文件的MD5值（十六进制字符串）
    """
    md5_hash = hashlib.md5()
    try:
        with open(file_path, 'rb') as f:
            # 分块读取文件，避免大文件占用过多内存
            for chunk in iter(lambda: f.read(4096), b''):
                md5_hash.update(chunk)
        return md5_hash.hexdigest()
    except Exception as e:
        print(f"计算MD5失败: {file_path}, 错误: {e}")
        return ""


def is_image_file(file_path: str) -> bool:
    """
    判断文件是否为图片
    
    Args:
        file_path: 文件路径
        
    Returns:
        是否为支持的图片格式
    """
    return Path(file_path).suffix.lower() in IMAGE_EXTENSIONS


def create_thumbnail(image_path: str, max_size: Tuple[int, int] = (150, 150)) -> str:
    """
    创建图片缩略图并保存到临时文件
    
    Args:
        image_path: 原始图片路径
        max_size: 缩略图最大尺寸 (宽, 高)
        
    Returns:
        临时缩略图文件路径，如果失败返回None
    """
    try:
        with Image.open(image_path) as img:
            # 转换为RGB模式（某些格式如RGBA需要转换才能保存为JPEG）
            if img.mode in ('RGBA', 'LA', 'P'):
                # 创建白色背景
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                img = background
            elif img.mode != 'RGB':
                img = img.convert('RGB')
            
            # 计算缩略图尺寸（保持宽高比）
            img.thumbnail(max_size, Image.Resampling.LANCZOS)
            
            # 保存到临时文件
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            img.save(temp_file.name, 'PNG')
            temp_file.close()
            return temp_file.name
    except Exception as e:
        print(f"  警告: 无法创建缩略图 {image_path}: {e}")
        return None


def scan_directory(directory: str) -> Dict[str, List[str]]:
    """
    扫描目录，计算所有文件的MD5值
    
    Args:
        directory: 要扫描的目录路径
        
    Returns:
        字典，key为MD5值，value为文件路径列表（可能多个文件有相同MD5）
    """
    md5_dict = {}
    directory = os.path.abspath(directory)
    
    print(f"\n正在扫描目录: {directory}")
    file_count = 0
    image_count = 0
    
    for root, dirs, files in os.walk(directory):
        for filename in files:
            file_path = os.path.join(root, filename)
            md5_value = calculate_md5(file_path)
            
            if md5_value:
                if md5_value not in md5_dict:
                    md5_dict[md5_value] = []
                md5_dict[md5_value].append(file_path)
                file_count += 1
                
                if is_image_file(file_path):
                    image_count += 1
                
                if file_count % 100 == 0:
                    print(f"  已处理 {file_count} 个文件（{image_count} 个图片）...")
    
    print(f"  完成！共处理 {file_count} 个文件，其中 {image_count} 个图片，{len(md5_dict)} 个唯一MD5值")
    return md5_dict


def compare_directories(dir1: str, dir2: str) -> List[Tuple[str, List[str], List[str]]]:
    """
    对比两个目录中MD5相同的文件
    
    Args:
        dir1: 第一个目录路径（国内版本）
        dir2: 第二个目录路径（国外版本）
        
    Returns:
        匹配结果列表，每个元素为 (md5值, dir1中的文件列表, dir2中的文件列表)
    """
    print("\n" + "="*60)
    print("开始扫描文件...")
    print("="*60)
    
    md5_dict1 = scan_directory(dir1)
    md5_dict2 = scan_directory(dir2)
    
    # 找到共同的MD5值
    common_md5 = set(md5_dict1.keys()) & set(md5_dict2.keys())
    
    print(f"\n找到 {len(common_md5)} 个MD5值匹配的文件组")
    
    # 构建结果列表
    results = []
    for md5_value in sorted(common_md5):
        results.append((md5_value, md5_dict1[md5_value], md5_dict2[md5_value]))
    
    return results


def get_relative_path(file_path: str, base_dir: str) -> str:
    """
    获取相对路径
    
    Args:
        file_path: 文件绝对路径
        base_dir: 基础目录
        
    Returns:
        相对路径
    """
    try:
        return os.path.relpath(file_path, base_dir)
    except:
        return file_path


def format_file_size(size_bytes: int) -> str:
    """
    格式化文件大小
    
    Args:
        size_bytes: 字节数
        
    Returns:
        格式化的文件大小字符串
    """
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024.0:
            return f"{size_bytes:.2f} {unit}"
        size_bytes /= 1024.0
    return f"{size_bytes:.2f} TB"


def export_to_excel(results: List[Tuple[str, List[str], List[str]]], 
                   output_path: str, 
                   dir1: str, 
                   dir2: str,
                   include_images: bool = True):
    """
    将对比结果导出到Excel文件
    
    Args:
        results: 对比结果列表
        output_path: 输出Excel文件路径
        dir1: 第一个目录路径
        dir2: 第二个目录路径
        include_images: 是否在Excel中插入图片预览
    """
    print(f"\n正在生成Excel文件: {output_path}")
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资源对比结果"
    
    # 设置表头样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入表头
    if include_images:
        headers = ["序号", "预览图", "MD5值", 
                   f"文件路径1 ({os.path.basename(dir1)})", 
                   f"文件路径2 ({os.path.basename(dir2)})", 
                   "文件大小", "文件类型"]
    else:
        headers = ["序号", "MD5值", 
                   f"文件路径1 ({os.path.basename(dir1)})", 
                   f"文件路径2 ({os.path.basename(dir2)})", 
                   "文件大小", "文件类型"]
    
    ws.append(headers)
    
    # 设置表头样式
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 临时文件列表（用于最后清理）
    temp_files = []
    
    # 写入数据
    row_num = 2
    image_count = 0
    
    for idx, (md5_value, files1, files2) in enumerate(results, 1):
        # 获取文件信息
        file1 = files1[0] if files1 else None
        file2 = files2[0] if files2 else None
        
        file_size = os.path.getsize(file1) if file1 else 0
        is_image = is_image_file(file1) if file1 else False
        file_type = "图片" if is_image else "其他"
        
        # 如果有多个文件对应，每个组合占一行
        max_count = max(len(files1), len(files2))
        
        for i in range(max_count):
            current_file1 = files1[i] if i < len(files1) else None
            current_file2 = files2[i] if i < len(files2) else None
            
            rel_path1 = get_relative_path(current_file1, dir1) if current_file1 else ""
            rel_path2 = get_relative_path(current_file2, dir2) if current_file2 else ""
            
            if include_images:
                row_data = [
                    idx if i == 0 else "",
                    "",  # 预览图占位
                    md5_value if i == 0 else "",
                    rel_path1,
                    rel_path2,
                    format_file_size(file_size) if i == 0 else "",
                    file_type if i == 0 else ""
                ]
            else:
                row_data = [
                    idx if i == 0 else "",
                    md5_value if i == 0 else "",
                    rel_path1,
                    rel_path2,
                    format_file_size(file_size) if i == 0 else "",
                    file_type if i == 0 else ""
                ]
            
            ws.append(row_data)
            
            # 设置对齐方式
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            # 插入图片（只插入一个预览图，因为MD5相同的文件内容完全一样）
            if include_images and is_image and i == 0:
                # 设置行高（以磅为单位，1磅≈0.353mm）
                ws.row_dimensions[row_num].height = 120
                
                # 优先使用第一个文件，如果不存在则使用第二个文件
                preview_file = None
                if current_file1 and os.path.exists(current_file1):
                    preview_file = current_file1
                elif current_file2 and os.path.exists(current_file2):
                    preview_file = current_file2
                
                # 插入预览图
                if preview_file:
                    thumb_path = create_thumbnail(preview_file, max_size=(150, 150))
                    if thumb_path:
                        temp_files.append(thumb_path)
                        try:
                            img = ExcelImage(thumb_path)
                            # 调整图片大小
                            img.width = 150
                            img.height = 150
                            # 插入到B列（预览图）
                            cell_pos = f'B{row_num}'
                            ws.add_image(img, cell_pos)
                            image_count += 1
                        except Exception as e:
                            print(f"  警告: 插入图片失败 {preview_file}: {e}")
            
            row_num += 1
            
            # 每处理50个文件显示一次进度
            if idx % 50 == 0:
                print(f"  已处理 {idx}/{len(results)} 组文件...")
    
    # 调整列宽
    if include_images:
        ws.column_dimensions['A'].width = 8      # 序号
        ws.column_dimensions['B'].width = 22     # 预览图
        ws.column_dimensions['C'].width = 35     # MD5
        ws.column_dimensions['D'].width = 50     # 路径1
        ws.column_dimensions['E'].width = 50     # 路径2
        ws.column_dimensions['F'].width = 12     # 文件大小
        ws.column_dimensions['G'].width = 10     # 文件类型
    else:
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 10
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存文件
    print(f"  正在保存Excel文件...")
    wb.save(output_path)
    
    # 清理临时文件
    print(f"  正在清理临时文件...")
    for temp_file in temp_files:
        try:
            if os.path.exists(temp_file):
                os.unlink(temp_file)
        except:
            pass
    
    print(f"Excel文件生成成功！")
    print(f"  - 共 {len(results)} 组匹配文件")
    if include_images:
        print(f"  - 已插入 {image_count} 个图片预览")


def main():
    """
    主函数
    """
    parser = argparse.ArgumentParser(
        description='游戏资源文件对比工具 - 通过MD5值匹配两个文件夹中的对应文件（支持图片预览）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  python compare_resources.py D:/assets/cn D:/assets/en D:/output/contrast.xlsx
  python compare_resources.py "C:/Game/Resources/CN" "C:/Game/Resources/EN" "C:/output/result.xlsx"
  python compare_resources.py D:/assets/cn D:/assets/en D:/output/result.xlsx --no-images
        """
    )
    
    parser.add_argument('dir1', type=str, help='第一个文件夹路径（国内版本）')
    parser.add_argument('dir2', type=str, help='第二个文件夹路径（国外版本）')
    parser.add_argument('output', type=str, help='输出Excel文件路径（例如: contrast.xlsx）')
    parser.add_argument('--no-images', action='store_true', 
                       help='不在Excel中插入图片预览（加快处理速度）')
    
    args = parser.parse_args()
    
    # 验证输入目录
    if not os.path.isdir(args.dir1):
        print(f"错误: 目录不存在: {args.dir1}")
        return
    
    if not os.path.isdir(args.dir2):
        print(f"错误: 目录不存在: {args.dir2}")
        return
    
    # 确保输出目录存在
    output_dir = os.path.dirname(args.output)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"创建输出目录: {output_dir}")
    
    # 确保输出文件有.xlsx扩展名
    if not args.output.lower().endswith('.xlsx'):
        args.output += '.xlsx'
    
    print("="*60)
    print("游戏资源文件对比工具（带图片预览）")
    print("="*60)
    print(f"目录1: {args.dir1}")
    print(f"目录2: {args.dir2}")
    print(f"输出文件: {args.output}")
    print(f"图片预览: {'否' if args.no_images else '是'}")
    
    # 执行对比
    results = compare_directories(args.dir1, args.dir2)
    
    # 导出到Excel
    if results:
        export_to_excel(results, args.output, args.dir1, args.dir2, 
                       include_images=not args.no_images)
        print("\n" + "="*60)
        print("处理完成！")
        print("="*60)
    else:
        print("\n警告: 没有找到任何匹配的文件")


if __name__ == '__main__':
    main()

